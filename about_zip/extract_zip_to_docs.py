# ===== Standard library =====
import os, re
import shutil
import zipfile
from pathlib import Path
from typing import List
from time import sleep

# def extract_zip_to_docs(zip_path: str, overwrite: bool = False) -> List[Path]:
#     """
#     C:\...\ZIP_74362\110-e\R2-2004593.zip を
#     C:\...\DOCS_74362\110-e\ に展開する。
#       - 単一ファイルZIP: DOCS側に  R2-2004593.<拡張子>
#       - 複数ファイルZIP: DOCS側に  R2-2004593_1.<拡張子>, R2-2004593_2.<拡張子>, ...
#     返り値: 作成/更新したファイルの Path リスト
#     """
#     p = Path(zip_path)
#     created: List[Path] = []

#     if not (p.is_file() and p.suffix.lower() == ".zip"):
#         print(f"⚠️ ZIPではありません: {p}")
#         return created

#     # 1) 祖先に ZIP* を探し、DOCS* を決定
#     zip_root = None
#     docs_root = None
#     for anc in p.parents:
#         name = anc.name
#         # ① ちょうど "ZIP"
#         if re.match(r'^zip$', name, re.I):
#             zip_root = anc
#             seq = ''
#             break
#         # ② "ZIP_74362" / "ZIP-74362"（数字のみサフィックス）
#         m = re.match(r'^zip[-_]?(\d+)$', name, re.I)
#         if m:
#             zip_root = anc
#             seq = m.group(1)
#             break
#         # ③ "ZIP_ai_is_6.1.3_kw_is_" など "ZIP_" で始まる任意サフィックス
#         m = re.match(r'^zip_(.+)$', name, re.I)
#         if m:
#             zip_root = anc
#             # ZIP_ の後ろをそのまま引き継いで DOCS_ に置換
#             seq = m.group(1)
#             break

#     if zip_root is None:
#         print(f"⚠️ ZIP 親フォルダが見つかりません（許容: 'ZIP', 'ZIP_…', 'ZIP-数字', 'ZIP_数字'）: {p}")
#         return created

#     # 2) DOCS_XXXX と相対サブフォルダ（例: 110-e）を決定
#     docs_root = zip_root.parent / f"DOCS_{seq}"
#     rel_subdir = p.parent.relative_to(zip_root)   # 例: '110-e'
#     dest_dir = docs_root / rel_subdir             # 例: DOCS_74362/110-e
#     dest_dir.mkdir(parents=True, exist_ok=True)

#     print(f"📂 ZIPルート: {zip_root.name} → 出力: {docs_root.name}")
#     print(f"🗜 解凍対象: {p.name}")

#     with zipfile.ZipFile(p) as zf:
#         members = [zi for zi in zf.infolist() if not zi.is_dir()]
#         if not members:
#             print(f"⚠️ 空のZIPです: {p}")
#             return created
#         members.sort(key=lambda z: z.filename.lower())

#         if len(members) == 1:
#             # 単一ファイル → zip名ベース + 中身拡張子
#             m = members[0]
#             inner_ext = Path(m.filename).suffix  # '.docx' 等 (無しなら空)
#             target = dest_dir / f"{p.stem}{inner_ext}"
#             created.append(target)
#             if target.exists() and not overwrite:
#                 print(f"↪ 既存のためスキップ: {target}")
#                 return created

#             target.parent.mkdir(parents=True, exist_ok=True)
#             with zf.open(m, 'r') as src, open(target, 'wb') as dst:
#                 shutil.copyfileobj(src, dst)
#             print(f"✅ 展開完了: {target}")
#             return created

#         else:
#             # 複数ファイル → ベース名_1, _2, ...
#             for idx, m in enumerate(members, start=1):
#                 ext = Path(m.filename).suffix
#                 target = dest_dir / f"{p.stem}_{idx}{ext}"
#                 created.append(target)
#                 if target.exists() and not overwrite:
#                     print(f"↪ 既存のためスキップ: {target}")
#                     continue
#                 target.parent.mkdir(parents=True, exist_ok=True)
#                 with zf.open(m, 'r') as src, open(target, 'wb') as dst:
#                     shutil.copyfileobj(src, dst)
#                 print(f"✅ 展開: {target}")

#             return created
#         return created
#     return created



def extract_zip_to_docs(zip_path: str, overwrite: bool = False) -> List[Path]:
    """
    C:\...\ZIP_74362\110-e\R2-2004593.zip を
    C:\...\DOCS_74362\110-e\ に展開する。
      - 単一ファイルZIP: DOCS側に  R2-2004593.<拡張子>
      - 複数ファイルZIP: DOCS側に  R2-2004593_1.<拡張子>, R2-2004593_2.<拡張子>, ...
    返り値: 作成/更新したファイルの Path リスト
    """
    p = Path(zip_path)
    dest_dir = p.parent
    created: List[Path] = []

    print(f"🗜 解凍対象: {p.name}")
    print(f"📂 出力先: {dest_dir}")

    with zipfile.ZipFile(p) as zf:
        # ディレクトリエントリを除外（中のサブフォルダはフラット化して書き出す）
        members = [zi for zi in zf.infolist() if not zi.is_dir()]
        if not members:
            print(f"⚠️ 空のZIPです: {p}")
            return created

        # 名前順に安定化
        members.sort(key=lambda z: z.filename.lower())

        if len(members) == 1:
            # 単一ファイル → zip名ベース + 中身拡張子
            m = members[0]
            inner_ext = Path(m.filename).suffix  # 例: '.docx'（無い場合は''）
            target = dest_dir / f"{p.stem}{inner_ext}"
            created.append(target)

            if target.exists() and not overwrite:
                print(f"↪ 既存のためスキップ: {target}")
                return created

            target.parent.mkdir(parents=True, exist_ok=True)
            with zf.open(m, 'r') as src, open(target, 'wb') as dst:
                shutil.copyfileobj(src, dst)

            print(f"✅ 展開完了: {target}")
            return created

        else:
            # 複数ファイル → ベース名_1, _2, ...
            for idx, m in enumerate(members, start=1):
                ext = Path(m.filename).suffix  # 各ファイルの拡張子を保持
                target = dest_dir / f"{p.stem}_{idx}{ext}"
                created.append(target)

                if target.exists() and not overwrite:
                    print(f"↪ 既存のためスキップ: {target}")
                    continue

                target.parent.mkdir(parents=True, exist_ok=True)
                with zf.open(m, 'r') as src, open(target, 'wb') as dst:
                    shutil.copyfileobj(src, dst)

                print(f"✅ 展開: {target}")

            return created

# pip install openpyxl
from pathlib import Path
from openpyxl import load_workbook

def extract_zip_to_docs_from_fold(download_path: str, file_name: str, overwrite: bool = False) -> List[List[Path]]:
    p = Path(download_path) / file_name
    res_zip: List[List[Path]] = []
    if p.suffix.lower() != ".xlsx":
        p = p.with_suffix(".xlsx")

    wb = load_workbook(p, read_only=True, data_only=True)
    ws = wb.active

    # 1行目（ヘッダー）を読み、saved_path列のインデックスを特定
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [("" if h is None else str(h).strip()) for h in header_row]

    saved_idx = None
    for i, h in enumerate(headers):
        if h.lower() == "saved_path":
            saved_idx = i
            break

    if saved_idx is None:
        print("saved_path 列が見つかりません。ヘッダー:", headers)
    else:
        # 2行目以降を走査して saved_path をすべてプリント
        for row in ws.iter_rows(min_row=2, values_only=True):
            v = row[saved_idx] if saved_idx < len(row) else None
            target = extract_zip_to_docs(v)
            res_zip.append(target)
    return res_zip

def clear_folder_files(folder_path: str):
    """
    指定されたフォルダ内の全ファイル・サブフォルダを削除する関数
    -----------------------------------------------------------------
    引数:
        folder_path (str): 削除対象のフォルダの絶対パス

    戻り値:
        なし
    """
    if not os.path.exists(folder_path):
        print(f"指定フォルダが存在しません: {folder_path}")
        return

    # フォルダ内を走査
    for item in os.listdir(folder_path):
        item_path = os.path.join(folder_path, item)

        try:
            if os.path.isfile(item_path) or os.path.islink(item_path):
                os.remove(item_path)
                print(f"🗑️ ファイル削除: {item_path}")
            elif os.path.isdir(item_path):
                shutil.rmtree(item_path)
                print(f"📁 フォルダ削除: {item_path}")
        except Exception as e:
            print(f"⚠️ 削除失敗: {item_path} ({e})")

    print(f"✅ フォルダ内容を全削除しました: {folder_path}")

def clean_text_edges(word: str) -> str:
    """
    文字列の前後にある空白・改行を削除する関数
    ------------------------------------------------
    引数:
        word (str): 対象の文字列

    戻り値:
        str: 前後の空白（半角・全角）および改行を削除した文字列
    """
    if not isinstance(word, str):
        return word  # 文字列以外はそのまま返す

    # 前後の半角/全角スペース、改行(\r, \n)を削除
    cleaned = re.sub(r'^[\s\u3000\r\n]+|[\s\u3000\r\n]+$', '', word)
    return cleaned