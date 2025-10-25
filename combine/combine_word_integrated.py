# combine_integrated_skip_bad.py
import os, sys, platform, subprocess
from openpyxl import load_workbook
import win32com.client as win32

def kill_all_word_processes():
    if platform.system().lower() != "windows":
        return
    try:
        subprocess.run(["taskkill", "/IM", "WINWORD.EXE", "/F", "/T"],
                       capture_output=True, text=True, check=False)
    except Exception:
        pass

def clean_document(doc):
    wdMainTextStory = 1
    # 事前クリーンアップ（変更履歴・コメント・H/F）
    try:
        if doc.Revisions.Count > 0:
            doc.AcceptAllRevisions()
    except Exception:
        pass
    try:
        if doc.Comments.Count > 0:
            for c in list(doc.Comments):
                try: c.Delete()
                except Exception: pass
    except Exception:
        pass
    for sec in doc.Sections:
        for idx in (1,2,3):  # Primary, First, Even
            try: sec.Headers(idx).Range.Delete()
            except Exception: pass
            try: sec.Footers(idx).Range.Delete()
            except Exception: pass
    # 全ストーリー削除（失敗したら本文だけでも空に）
    try:
        doc.Content.Delete()
    except Exception:
        pass
    doc.Content.InsertAfter("\r")  # 空段落保証

def is_file_healthy(word_app, path: str) -> bool:
    """OpenAndRepair で開けるかをチェック。開けなければ破損とみなす。"""
    doc = None
    try:
        # ConfirmConversions=False, ReadOnly=True, OpenAndRepair=True
        doc = word_app.Documents.Open(path, False, True, None, None, None, None, None, None, None, True)
        return True
    except Exception:
        return False
    finally:
        try:
            if doc is not None:
                doc.Close(False)
        except Exception:
            pass


def combine_word_integrated(
    excel_path: str,
    *,
    with_markers: bool = True,
    kill_word: bool = True,
    accept_revisions: bool = True,
    remove_comments: bool = True,
):
    """
    破損判定を緩くした版：
    - 事前の「健康診断」は行わない
    - まず tmp_doc.Range.InsertFile(path) を試す
    - 失敗したら Documents.Open(..., OpenAndRepair=True) で開いて StoryRanges を転送
    - それもダメならスキップ
    """
    if kill_word:
        kill_all_word_processes()

    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    output_path = ws["B1"].value
    if not output_path or not isinstance(output_path, str):
        raise ValueError("B1 に出力先（フォルダ or .docx）を指定してください。")
    if not output_path.lower().endswith(".docx"):
        output_path = os.path.join(output_path, "combined_all.docx")

    valid_exts = (".doc", ".docx", ".docm")
    files = []
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        p = row[0].value
        if p and isinstance(p, str) and os.path.isfile(p) and p.lower().endswith(valid_exts):
            files.append(p)
    if not files:
        print("❌ 入力ファイルが見つかりません。")
        return

    skipped = []

    word = None
    merged_doc = None
    try:
        word = win32.DispatchEx("Word.Application")
        try:
            word.Visible = False
            word.ScreenUpdating = False
            word.DisplayAlerts = 0
            word.Options.CheckSpellingAsYouType = False
            word.Options.CheckGrammarAsYouType = False
        except Exception:
            pass

        merged_doc = word.Documents.Add()
        clean_document(merged_doc)

        # 定数
        wdPageBreak = 7
        wdMainTextStory = 1
        wdHeaderFooterIndex = (1, 2, 3)
        wdCollapseEnd = 0

        def end_range():
            rng = merged_doc.Content.Duplicate
            rng.Collapse(Direction=wdCollapseEnd)
            return rng

        for i, path in enumerate(files, start=1):
            name = os.path.basename(path)
            print(f"⚡ {i}/{len(files)}: {name}")

            tmp_doc = word.Documents.Add()
            try:
                clean_document(tmp_doc)

                inserted = False
                # ① まずは高速経路：InsertFile を試す
                try:
                    tmp_doc.Range(0, 0).InsertFile(path)
                    inserted = True
                except Exception:
                    inserted = False

                # ② InsertFile がダメなら、OpenAndRepair で開いて本文を転送
                if not inserted:
                    src_doc = None
                    try:
                        # ConfirmConversions=False, ReadOnly=True, AddToRecentFiles=False, PasswordDocument=None, OpenAndRepair=True
                        src_doc = word.Documents.Open(path, False, True, None, None, None, None, None, None, None, True)
                        # 変更履歴/コメントはここでも処理できる
                        if accept_revisions:
                            try:
                                if src_doc.Revisions.Count > 0:
                                    src_doc.AcceptAllRevisions()
                            except Exception:
                                pass
                        if remove_comments:
                            try:
                                if src_doc.Comments.Count > 0:
                                    for c in list(src_doc.Comments):
                                        try: c.Delete()
                                        except Exception: pass
                            except Exception:
                                pass
                        # ヘッダー/フッター除去
                        for sec in src_doc.Sections:
                            for idx in wdHeaderFooterIndex:
                                try: sec.Headers(idx).Range.Delete()
                                except Exception: pass
                                try: sec.Footers(idx).Range.Delete()
                                except Exception: pass
                        # 本文を tmp_doc へ直接流し込む（書式ごと）
                        try:
                            main_rng = src_doc.StoryRanges(wdMainTextStory)
                        except Exception:
                            main_rng = None
                        if main_rng is not None and getattr(main_rng, "StoryLength", 0) > 0:
                            tmp_doc.Range(0, 0).FormattedText = main_rng.FormattedText
                            inserted = True
                    except Exception:
                        inserted = False
                    finally:
                        try:
                            if src_doc is not None:
                                src_doc.Close(False)
                        except Exception:
                            pass

                if not inserted:
                    print(f"⏭️ 読み込み不能のためスキップ: {name}")
                    skipped.append(path)
                    continue

                # --- 以降は従来通り（tmp_doc 内で整形してから結合） ---

                # 変更履歴承諾（InsertFile経路で残っていた場合）
                if accept_revisions:
                    try:
                        if tmp_doc.Revisions.Count > 0:
                            tmp_doc.AcceptAllRevisions()
                    except Exception:
                        pass

                # コメント削除
                if remove_comments:
                    try:
                        if tmp_doc.Comments.Count > 0:
                            for c in list(tmp_doc.Comments):
                                try: c.Delete()
                                except Exception: pass
                    except Exception:
                        pass

                # ヘッダー/フッター削除（全セクション）
                for sec in tmp_doc.Sections:
                    for idx in wdHeaderFooterIndex:
                        try: sec.Headers(idx).Range.Delete()
                        except Exception: pass
                        try: sec.Footers(idx).Range.Delete()
                        except Exception: pass

                # マーカー/改ページは tmp_doc 側に埋める
                if with_markers:
                    try:
                        tmp_doc.Content.InsertBefore(f"\r\n#_C#O#P#Y_# {name} のコピー開始\r\n\r\n")
                    except Exception:
                        pass
                    try:
                        tmp_doc.Content.InsertAfter("\r\n\r\n")
                        r_end = tmp_doc.Content.Duplicate; r_end.Collapse(Direction=wdCollapseEnd)
                        r_end.InsertAfter(f"## {name} のコピー終了\r\n")
                    except Exception:
                        pass
                if i < len(files):
                    try:
                        tmp_doc.Content.InsertAfter("\r")
                        r_end = tmp_doc.Content.Duplicate; r_end.Collapse(Direction=wdCollapseEnd)
                        r_end.InsertBreak(Type=wdPageBreak)
                    except Exception:
                        pass

                # 本文(MainTextStory) を結合先へ 1 回だけ転送
                try:
                    src_main = tmp_doc.StoryRanges(wdMainTextStory)
                except Exception:
                    src_main = None
                if src_main is None or getattr(src_main, "StoryLength", 0) == 0:
                    print(f"⏭️ 本文が空のためスキップ: {name}")
                    skipped.append(path)
                    continue

                try:
                    end_range().FormattedText = src_main.FormattedText
                except Exception:
                    print(f"⏭️ 貼り付け失敗のためスキップ: {name}")
                    skipped.append(path)
                    continue

            finally:
                try:
                    tmp_doc.Close(False)
                except Exception:
                    pass

        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        merged_doc.SaveAs2(output_path)
        print(f"✅ 完全統合完了: {output_path}")

        if skipped:
            print("⚠️ スキップしたファイル（最終的に処理不能）:")
            for s in skipped:
                print("  -", s)

    except Exception:
        try:
            if merged_doc is not None:
                base, ext = os.path.splitext(output_path)
                merged_doc.SaveAs2(f"{base}__PARTIAL{ext}")
                print(f"⚠️ 途中までを保存しました: {base}__PARTIAL{ext}")
        except Exception:
            pass
        raise
    finally:
        try:
            if merged_doc is not None:
                merged_doc.Close(False)
        except Exception:
            pass
        try:
            if word is not None:
                try:
                    word.ScreenUpdating = True
                except Exception:
                    pass
                word.Quit()
        except Exception:
            pass
        kill_all_word_processes()
