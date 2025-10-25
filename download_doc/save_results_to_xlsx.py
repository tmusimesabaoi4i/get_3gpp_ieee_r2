# pip install pandas openpyxl
from pathlib import Path
from typing import Any, Iterable
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ===== 出力列の順序 =====
COLUMNS = ["index", "url", "filename", "download_path", "name", "saved_path", "ext", "skipped"]

def _ensure_xlsx_name(name: str | Path) -> str:
    """拡張子 .xlsx を保証して返す"""
    s = str(name)
    return s if s.lower().endswith(".xsx") else s + ".xlsx"

def save_results_to_xlsx(results: Iterable[dict[str, Any]],
                         folder: str | Path,
                         filename: str | Path) -> Path:
    """
    results（辞書のリスト等）を folder/filename(.xlsx) に保存する。
    - 列順は COLUMNS に固定
    - 文字列は trim と CRLF→LF 正規化
    - オートフィルタ、ヘッダー固定、列幅自動調整
    """
    # ===== DataFrame 構築（列順固定） =====
    rows = []
    for r in results:
        row = {}
        for col in COLUMNS:
            v = r.get(col, None)
            if isinstance(v, (str, Path)):
                s = str(v).replace("\r\n", "\n").strip()
                row[col] = s
            else:
                row[col] = v
        rows.append(row)
    df = pd.DataFrame(rows, columns=COLUMNS)

    # ===== パス決定 =====
    folder = Path(folder)
    folder.mkdir(parents=True, exist_ok=True)
    xlsx_path = folder / _ensure_xlsx_name(filename)

    # ===== 保存 =====
    df.to_excel(xlsx_path, index=False)

    # ===== 見た目調整 =====
    wb = load_workbook(xlsx_path)
    ws = wb.active

    # オートフィルタ（全範囲）
    last_col_letter = get_column_letter(ws.max_column)
    last_row = ws.max_row
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    # ヘッダー固定
    ws.freeze_panes = "A2"

    # 列幅自動調整
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        max_len = len(col_name)
        for cell_vals in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, values_only=True):
            for v in cell_vals:
                if v is None:
                    continue
                s = str(v)
                max_len = max(max_len, len(s))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 80)

    wb.save(xlsx_path)
    return xlsx_path

# pip install pandas openpyxl
from pathlib import Path
from typing import Iterable, List, Union
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def write_res_zip_paths_to_xlsx(
    res_zip: Union[List[List[Path]], List[Path], Iterable],
    folder: Union[str, Path],
    filename: Union[str, Path],
) -> Path:
    """
    res_zip 内の全 Path を 1列 (saved_path) にして folder/filename(.xlsx) に保存する。

    - res_zip は List[List[Path]]（二重リスト）を想定するが、一次元（List[Path]）や
      文字列混在でも Path に変換できれば取り込み可。
    - 文字列は前後の空白と改行を除去（CRLF→LF→strip）
    - A1 にヘッダー 'saved_path'、オートフィルタ、ヘッダー固定(A2)、列幅自動調整
    """
    # ===== パス展開（フラット化） =====
    flat_paths: List[str] = []

    def _push(v):
        if v is None:
            return
        s = str(Path(v))  # Path化して正規化
        s = s.replace("\r\n", "\n").strip()
        flat_paths.append(s)

    for item in res_zip:
        # 二重/一次元どちらも受ける
        if isinstance(item, (list, tuple, set)):
            for p in item:
                _push(p)
        else:
            _push(item)

    # ===== DataFrame 作成 =====
    df = pd.DataFrame({"saved_path": flat_paths})

    # ===== 出力先決定 =====
    folder = Path(folder)
    folder.mkdir(parents=True, exist_ok=True)
    filename = str(filename)
    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"
    xlsx_path = folder / filename

    # ===== 書き出し =====
    df.to_excel(xlsx_path, index=False)

    # ===== 見た目調整 =====
    wb = load_workbook(xlsx_path)
    ws = wb.active

    # オートフィルタ（全範囲）
    last_col_letter = get_column_letter(ws.max_column)
    last_row = ws.max_row
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    # ヘッダー固定（1行目固定 → 表示開始を2行目に）
    ws.freeze_panes = "A2"

    # 列幅自動調整（最大長 + 余白）
    max_len = len("saved_path")
    for col_vals in ws.iter_cols(min_col=1, max_col=1, min_row=2, values_only=True):
        for v in col_vals:
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
    ws.column_dimensions["A"].width = min(max_len + 2, 80)

    wb.save(xlsx_path)
    return xlsx_path

# pip install openpyxl
from pathlib import Path
from typing import List, Any, Union
from openpyxl import load_workbook

def read_column_as_list(folder: Union[str, Path],
                        filename: Union[str, Path],
                        col_index: int) -> List[Any]:
    """
    指定列（0=A, 1=B, ...）の値をリストで返す。1行目（ヘッダー）は除外。
    返却値はセルの値（数値/日付/文字列など openpyxl の解釈結果）。
    """
    p = Path(folder) / filename
    if p.suffix.lower() != ".xlsx":
        p = p.with_suffix(".xlsx")

    wb = load_workbook(p, read_only=True, data_only=True)
    ws = wb.active

    values: List[Any] = []
    for (v,) in ws.iter_rows(
        min_row=2,                  # 2行目から（ヘッダー除外）
        min_col=col_index + 1,     # openpyxlは1始まり
        max_col=col_index + 1,
        values_only=True
    ):
        values.append(v)
    return values

# 使い方例:
# lst = read_column_as_list(r"C:\Users\yohei\Downloads", "outfile.xlsx", 0)  # A列
# print(lst)


# ===== 使い方例 =====
if __name__ == "__main__":
    results = [
        {
            "index": 1,
            "row": "R2-2206906",
            "url": "http://example.com/file1.zip",
            "filename": "R2-2206906_file1.zip",
            "saved_path": r"C:\Users\yohei\Downloads\R2-2206906\R2-2206906_file1.zip",
            "ext": ".zip",
            "skipped": True,
        }
    ]
    out = save_results_to_xlsx(results, r"C:\Users\yohei\Downloads", "download_results")
    print("✅ 書き出し:", out)
