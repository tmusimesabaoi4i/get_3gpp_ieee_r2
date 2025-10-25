#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
main.py — 完全統合:
- 引数は Excel 絶対パスのみ（引用符付き想定）
- 事前/事後で Office Kill（psutil→taskkill）
- Ctrl+C/SIGTERM で中断
- ExcelReader: 列/行/セルを最速志向で取得（すべて str で返却、端の空白/改行/特殊文字 trim）
"""

from __future__ import annotations

import argparse
import logging
import signal
import subprocess
import sys
import re
from pathlib import Path
from typing import Any, List, Optional, Literal, Dict, Tuple, Callable








from util import (
    get_proxy_from_cmd,
    get_downloads_path,
    build_case_folder_from_excel,
)


from download_doc.download_doc_3gpp import(
    # fetch_3gpp_docs,
    fetch_3gpp_docs_queue
)

from download_doc.download_doc_ieee import(
    # fetch_ieee_docs,
    fetch_ieee_docs_queue
)

from download_doc.save_results_to_xlsx import(
    save_results_to_xlsx,
    write_res_zip_paths_to_xlsx,
    read_column_as_list
)

from about_zip.extract_zip_to_docs import(
    extract_zip_to_docs_from_fold
)

from combine.extract_paragraphs import(
    convert_office_to_html
)











# ============== アプリ基本情報 ==============
__app_name__ = "myapp"
__version__ = "0.4.0"

# ============== ログ設定 ==============
def setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s [%(name)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

# ============== Office Kill ==============
OFFICE_PROCESSES: tuple[str, ...] = (
    "WINWORD.EXE",
    "EXCEL.EXE",
    "POWERPNT.EXE",
    "MSACCESS.EXE",
    "ONENOTE.EXE",
    "OUTLOOK.EXE",
    "VISIO.EXE",
    "WINPROJ.EXE",
    "LYNC.EXE",
    "MSPUB.EXE",
    "MSOSYNC.EXE",
    "OFFICECLICKTORUN.EXE",
)

def _kill_by_psutil(names: Iterable[str]) -> bool:
    try:
        import psutil  # type: ignore
    except Exception:
        return False
    log = logging.getLogger(f"{__app_name__}.kill.psutil")
    targets = {n.lower() for n in names}
    procs = []
    for p in psutil.process_iter(attrs=["pid", "name"]):
        try:
            name = (p.info.get("name") or "").lower()
            if name in targets:
                log.info("terminate pid=%s name=%s", p.pid, name)
                p.terminate()
                procs.append(p)
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            continue
    if procs:
        gone, alive = psutil.wait_procs(procs, timeout=8.0)
        for p in alive:
            try:
                log.warning("force kill pid=%s name=%s", p.pid, p.name())
                p.kill()
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                continue
    return True

def _kill_by_taskkill(names: Iterable[str]) -> None:
    log = logging.getLogger(f"{__app_name__}.kill.taskkill")
    for n in names:
        try:
            res = subprocess.run(["taskkill", "/F", "/T", "/IM", n],
                                 capture_output=True, text=True, timeout=10)
            if res.stdout:
                log.debug("[taskkill stdout] %s", res.stdout.strip())
            if res.stderr:
                log.debug("[taskkill stderr] %s", res.stderr.strip())
        except subprocess.TimeoutExpired:
            log.error("taskkill timeout: %s", n)
        except FileNotFoundError:
            log.error("taskkill not found (Windows専用)")

def kill_office_processes() -> None:
    if not _kill_by_psutil(OFFICE_PROCESSES):
        logging.getLogger(__app_name__).debug("psutil unavailable; fallback to taskkill")
        _kill_by_taskkill(OFFICE_PROCESSES)

# ============== 文字列サニタイズ（共通） ==============
def sanitize_to_str(v: Any) -> str:
    """
    値を str 化し、前後の空白/改行/特殊文字を trim:
      - None → ""
      - 改行(\r\n, \r, \n)は空白1個に置換
      - ゼロ幅/制御文字(\u0000-\u001F,\u007F-\u009F, \u200B-\u200D, \uFEFF)を除去
      - 前後の空白（半角/全角 \u3000）を strip
    """
    if v is None:
        s = ""
    else:
        s = str(v)
    if not s:
        return ""

    # 改行 → 半角スペース
    s = s.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")

    # ゼロ幅/制御文字の除去
    # 注意: 内部の通常スペースは残す（要件は前後trim）
    import re
    s = re.sub(r"[\u0000-\u001F\u007F-\u009F\u200B-\u200D\uFEFF]", "", s)

    # 前後の空白（半角/全角）を除去
    s = s.strip(" \t\u3000")
    return s

# ============== Excel Reader（高速ランダムアクセス） ==============
Mode = Literal["row", "col", "cell"]

_HYPERLINK_RE = re.compile(
    r'^\s*=\s*HYPERLINK\s*\(\s*("([^"]+)"|\'([^\']+)\')\s*,', re.IGNORECASE
)

def sanitize_to_str(v: Any) -> str:
    # 既存の sanitize_to_str が別にあるなら置き換えてOK
    if v is None:
        return ""
    return str(v).strip()

def _fast_hl_from_cell(cell) -> str:
    """セルのリンク先だけを最小コストで抽出。"""
    hl = getattr(cell, "hyperlink", None)
    if hl:
        tgt = getattr(hl, "target", None)
        if isinstance(hl, str):
            return hl
        if isinstance(tgt, str):
            return tgt
    # =HYPERLINK() の数式（data_only=False で取得）
    if getattr(cell, "data_type", None) == "f":
        v = cell.value
        if isinstance(v, str):
            m = _HYPERLINK_RE.match(v)
            if m:
                return m.group(2) or m.group(3) or ""
    return ""

def _iter_hyperlink_coords(ws) -> Iterable[Tuple[int, int]]:
    """
    シート内の『明示リンク（関係付き）』が指す全セル座標 (row, col) を列挙。
    =HYPERLINK() 数式はここには含まれない（セル側で解析）。
    """
    from openpyxl.utils.cell import range_boundaries
    for hl in getattr(ws, "_hyperlinks", []):  # 内部配列。速い
        ref = getattr(hl, "ref", None)
        if not ref:
            continue
        min_c, min_r, max_c, max_r = range_boundaries(ref)
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                yield (r, c)

class ExcelReader:
    """
    “指定された場所だけ読む” 高速版（ノーキャッシュ）。
    - 列/行の末端は、実セル辞書（ws._cells）＋明示リンク範囲だけを見て決定
    - 読み出しはその最小範囲のみ iter_rows()
    - hyperlink=True のときは、その範囲内セルのみで URL を抽出
    """
    def __init__(
        self,
        excel_path: Path,
        *,
        read_only: bool = False,   # 明示リンクのみで良いなら True が速い（=HYPERLINK() を使うなら False 推奨）
    ):
        self.path = Path(excel_path)
        if not self.path.is_file():
            raise FileNotFoundError(self.path)
        import openpyxl
        # =HYPERLINK() 解析には data_only=False が必要
        self._wb = openpyxl.load_workbook(
            self.path,
            data_only=False,
            read_only=read_only,
            keep_links=True,  # 明示リンクの解決
        )

    # ---- 基本 ----
    def _get_ws(self, sheet: Optional[str | int] = None):
        if sheet is None:
            return self._wb.active
        if isinstance(sheet, int):
            names = self._wb.sheetnames
            if not (1 <= sheet <= len(names)):
                raise IndexError(f"sheet index out of range: {sheet}")
            return self._wb[names[sheet - 1]]
        return self._wb[sheet]

    # ---- 末端検出（列/行ごとに最小範囲）----
    def _last_row_in_col(self, ws, col: int, min_row: int) -> int:
        """
        その列に『値 or 数式 or 明示リンク』が存在する最終行を返す。
        無ければ min_row-1（=空）。
        """
        cells = getattr(ws, "_cells", {})  # 実セルのみ。空は入らない
        last = 0
        if cells:
            # keys() は (row, col) タプル
            for (r, c) in cells.keys():
                if c == col and r >= min_row and r > last:
                    last = r
        # 明示リンクが指すセルも含める
        for (r, c) in _iter_hyperlink_coords(ws):
            if c == col and r >= min_row and r > last:
                last = r
        return last if last >= min_row else (min_row - 1)

    def _last_col_in_row(self, ws, row: int, min_col: int) -> int:
        """
        その行に『値 or 数式 or 明示リンク』が存在する最終列を返す。
        無ければ min_col-1。
        """
        cells = getattr(ws, "_cells", {})
        last = 0
        if cells:
            for (r, c) in cells.keys():
                if r == row and c >= min_col and c > last:
                    last = c
        for (r, c) in _iter_hyperlink_coords(ws):
            if r == row and c >= min_col and c > last:
                last = c
        return last if last >= min_col else (min_col - 1)

    # ---- コア読取（範囲だけ iter_rows）----
    def _read_col_core(self, ws, n: int, start_row: int, hyperlink: bool) -> List[str]:
        if n < 1:
            raise IndexError("column index must be >= 1")
        last = self._last_row_in_col(ws, n, start_row)
        if last < start_row:
            return []
        # 1列だけの iter_rows は高速
        out: List[str] = []
        for (cell,) in ws.iter_rows(min_col=n, max_col=n, min_row=start_row, max_row=last):
            out.append(_fast_hl_from_cell(cell) if hyperlink else sanitize_to_str(cell.value))
        return out

    def _read_row_core(self, ws, n: int, start_col: int, hyperlink: bool) -> List[str]:
        if n < 1:
            raise IndexError("row index must be >= 1")
        last = self._last_col_in_row(ws, n, start_col)
        if last < start_col:
            return []
        out: List[str] = []
        for row in ws.iter_rows(min_row=n, max_row=n, min_col=start_col, max_col=last):
            for cell in row:
                out.append(_fast_hl_from_cell(cell) if hyperlink else sanitize_to_str(cell.value))
        return out

    # ---- 公開API（互換）----
    def read_col(
        self,
        n: int,
        *,
        sheet: Optional[str | int] = None,
        header: bool = False,
        hyperlink: bool = False,
    ) -> List[str]:
        ws = self._get_ws(sheet)
        start_row = 2 if header else 1
        return self._read_col_core(ws, n, start_row, hyperlink)

    def read_row(
        self,
        n: int,
        *,
        sheet: Optional[str | int] = None,
        header: bool = False,
        hyperlink: bool = False,
    ) -> List[str]:
        ws = self._get_ws(sheet)
        start_col = 2 if header else 1
        return self._read_row_core(ws, n, start_col, hyperlink)

    def read_cell(
        self,
        i: int,
        j: int,
        *,
        sheet: Optional[str | int] = None,
        hyperlink: bool = False,
    ) -> str:
        if i < 1 or j < 1:
            raise IndexError("row/column index must be >= 1")
        ws = self._get_ws(sheet)
        cell = ws.cell(row=i, column=j)
        return _fast_hl_from_cell(cell) if hyperlink else sanitize_to_str(cell.value)

    def xread(
        self,
        mode: Mode,
        a: int,
        b: Optional[int] = None,
        *,
        sheet: Optional[str | int] = None,
        header: bool = False,
        hyperlink: bool = False,
    ) -> Any:
        if mode == "col":
            return self.read_col(a, sheet=sheet, header=header, hyperlink=hyperlink)
        if mode == "row":
            return self.read_row(a, sheet=sheet, header=header, hyperlink=hyperlink)
        if mode == "cell":
            if b is None:
                raise ValueError("xread('cell', i, j=...) requires both i and j")
            return self.read_cell(a, b, sheet=sheet, hyperlink=hyperlink)
        raise ValueError(f"unknown mode: {mode}")

    def close(self) -> None:
        try:
            self._wb.close()
        except Exception:
            pass

# ============== 引数処理（Excel絶対パスのみ） ==============
def normalize_quoted_path(s: str) -> Path:
    """
    引用符付きの絶対パスを正規化。
    - 半角/全角の各種引用符を除去
    """
    s = s.strip()
    QUOTES = '"\'“”‘’«»「」『』‹›'
    while s and s[0] in QUOTES:
        s = s[1:]
    while s and s[-1] in QUOTES:
        s = s[:-1]
    return Path(s)

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog=__app_name__,
        description="Run with a single Excel absolute path. Office processes are killed before/after.",
        add_help=True,
    )
    p.add_argument("--version", action="version", version=f"%(prog)s {__version__}")
    p.add_argument("excel_path", help="Excelファイルの絶対パス（必ず引用符付きで指定）")
    return p

# ============== メイン処理 ==============
_SHOULD_STOP = False
def _signal_handler(signum, frame):
    global _SHOULD_STOP
    logging.getLogger(__app_name__).warning("received signal %s; shutting down...", signum)
    _SHOULD_STOP = True

def run(excel_path: Path) -> int:
    """
    ここに実処理を実装。例として print('process')。
    必要に応じて ExcelReader を用いて設定値を取得してください。
    """
    # 例: 最小動作（ここを置き換えて拡張）
    xr = ExcelReader(excel_path)

    sheet_initial_setup = int(xr.xread("cell", 10, 4, sheet = 1))
    sheet_url_list_path = int(xr.xread("cell", 11, 4, sheet = sheet_initial_setup))

    sheet_search_targets_3gpp = int(xr.xread("cell", 12, 4, sheet = sheet_initial_setup))
    sheet_search_targets_ieee = int(xr.xread("cell", 13, 4, sheet = sheet_initial_setup))
    sheet_search_rules_3gpp = int(xr.xread("cell", 14, 4, sheet = sheet_initial_setup))
    sheet_search_rules_ieee = int(xr.xread("cell", 15, 4, sheet = sheet_initial_setup))


    proxy_url = xr.xread("cell", 1, 3, sheet = sheet_initial_setup)
    download_dir = xr.xread("cell", 2, 3, sheet = sheet_initial_setup)
    database = xr.xread("cell", 3, 3, sheet = sheet_initial_setup)
    html_dir = xr.xread("cell", 4, 3, sheet = sheet_initial_setup)
    xlsx_dir = xr.xread("cell", 5, 3, sheet = sheet_initial_setup)
    zip_dir = xr.xread("cell", 6, 3, sheet = sheet_initial_setup)
    doc_dir = xr.xread("cell", 7, 3, sheet = sheet_initial_setup)
    combined_html_dir = xr.xread("cell", 8, 3, sheet = sheet_initial_setup)

    hyperlink_3gppp = xr.xread("cell", 1, 10, sheet = sheet_initial_setup)
    hyperlink_ieee = xr.xread("cell", 2, 10, sheet = sheet_initial_setup)


    if str(database) == "3gpp":
        print("3gpp")
        download_urls =  xr.xread("col", 1, header=True, sheet = sheet_url_list_path, hyperlink = hyperlink_3gppp)
        res = fetch_3gpp_docs_queue(str(download_dir),download_urls,proxy=proxy_url)
        save_results_to_xlsx(res,str(download_dir),"out_"+str(database))
        res_zip = extract_zip_to_docs_from_fold(str(download_dir),"out_"+str(database))
        write_res_zip_paths_to_xlsx(res_zip,str(download_dir),"out_file_"+str(database))
        l = read_column_as_list(str(download_dir),"out_file_"+str(database),0)
        convert_office_to_html(l,Path(download_dir) / "combine.html")

    if str(database) == "ieee":
        print("ieee")
        download_urls =  xr.xread("col", 1, header=True, sheet = sheet_url_list_path, hyperlink = hyperlink_ieee)
        res = fetch_ieee_docs_queue(str(download_dir),download_urls,proxy=proxy_url)
        save_results_to_xlsx(res,str(download_dir),"out_"+str(database))
        l = read_column_as_list(str(download_dir),"out_"+str(database),5)
        convert_office_to_html(l,Path(download_dir) / "combine.html")

    # --- 例: 設定値の読み取り（必要ならコメントアウト解除） ---
    # xr = ExcelReader(excel_path)
    # first_row = xr.xread("row", 1, header=False)    # 1行目をリスト[str]で
    # fifth_col = xr.xread("col", 5, header=True)     # 5列目、1行目は見出しなのでスキップ
    # v_10_2   = xr.xread("cell", 10, 2)              # (10行,2列) を単一の str で
    # logging.getLogger(__app_name__).info("row1=%s, col5(len)=%d, (10,2)=%s",
    #                                      first_row, len(fifth_col), v_10_2)

    logging.getLogger(__app_name__).info("run() executed with: %s", excel_path)
    return 0

def main(argv: list[str] | None = None) -> int:
    if argv is None:
        argv = sys.argv[1:]

    setup_logging()

    # Ctrl+C / SIGTERM
    signal.signal(signal.SIGINT, _signal_handler)
    try:
        signal.signal(signal.SIGTERM, _signal_handler)
    except Exception:
        pass

    parser = build_parser()
    args, unknown = parser.parse_known_args(argv)

    # PyInstaller の mp 子プロセスが付けるフラグだけ許容して捨てる
    allowed_mp_flags = tuple(["--multiprocessing", "--forkserver", "--spawn", "--freeze", "pipe_handle="])
    noise = [u for u in unknown if any(k in u for k in allowed_mp_flags)]
    others = [u for u in unknown if u not in noise]
    if others:
        parser.error(f"unrecognized arguments: {' '.join(others)}")


    excel_path = normalize_quoted_path(args.excel_path)
    if not excel_path.is_file():
        logging.getLogger(__app_name__).error("Excelファイルが見つかりません: %s", excel_path)
        return 2
    if excel_path.suffix.lower() not in {".xlsx", ".xlsm", ".xls"}:
        logging.getLogger(__app_name__).warning("Excel拡張子ではない可能性があります: %s", excel_path)

    # ---- Before: Office Kill ----
    logging.getLogger(__app_name__).info("killing Office processes (before)")
    kill_office_processes()
    if _SHOULD_STOP:
        logging.getLogger(__app_name__).info("interrupted before main run")
        return 130

    try:
        rc = run(excel_path)
    except KeyboardInterrupt:
        logging.getLogger(__app_name__).info("interrupted by user")
        rc = 130
    except Exception:
        logging.getLogger(__app_name__).exception("unhandled exception in run()")
        rc = 1

    # ---- After: Office Kill ----
    logging.getLogger(__app_name__).info("killing Office processes (after)")
    kill_office_processes()

    return rc

if __name__ == "__main__":
    import multiprocessing as mp
    mp.freeze_support()   # ★ 凍結バイナリでは必須
    raise SystemExit(main())